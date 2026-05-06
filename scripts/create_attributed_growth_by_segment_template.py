#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import re
import tempfile
from copy import deepcopy
from pathlib import Path
from typing import Any

try:
    import xlsxwriter
except ModuleNotFoundError as exc:
    raise SystemExit(
        "Missing dependency: XlsxWriter. Run: python3 -m venv .venv && "
        ".venv/bin/pip install -r requirements.txt"
    ) from exc

try:
    from PIL import Image
except ModuleNotFoundError:
    Image = None

SCRIPT_DIR = Path(__file__).resolve().parent
REPO_ROOT = SCRIPT_DIR.parent
DEFAULT_BANK_ORDER = ["RBC", "BMO", "Scotia", "CIBC", "National Bank", "TD"]
DEFAULT_INPUT_PATH = (
    REPO_ROOT
    / "inputs"
    / "attributed-growth-by-segment-template"
    / "attributed_growth_by_segment_template_input.xlsx"
)
DEFAULT_OUTPUT_PATH = (
    REPO_ROOT
    / "outputs"
    / "attributed-growth-by-segment-template"
    / "attributed_growth_by_segment_template.xlsx"
)
DEFAULT_LOGO_DIR = REPO_ROOT / "assets" / "bank-logos"
DEFAULT_LOGO_FILES = {
    "RBC": "ry.png",
    "BMO": "bmo.png",
    "Scotia": "bns.png",
    "CIBC": "cm.png",
    "National Bank": "na.png",
    "TD": "td.png",
}
CENTER_COLS = [2, 4, 6, 8, 10, 12]
LAYOUT = {
    "absolute_label_row": 3,
    "total_label_row": 4,
    "chart_top_row": 5,
    "chart_top_col": 0,
    "logo_top_row": 29,
    "logo_bottom_row": 30,
    "source_table_start_row": 3,
}
CHART_GEOMETRY = {
    "top_row": 5,
    "bottom_row": 29,
    "bar_width_px": 92,
    "label_width_px": 54,
    "label_height_px": 21,
    "axis_max_padding": 1.03,
    "bar_overlap_px": 1.0,
    "zero_marker_height_px": 4,
    "zero_line_color": "#8C8C8C",
}
INPUT_LAYOUT = {
    "sheet_name": "Data Input",
    "segment_start_row": 5,
    "segment_max_rows": 5,
    "data_start_row": 12,
    "data_max_rows": 6,
    "data_segment_start_col": 3,
}
COLUMN_WIDTHS_PX = {
    0: 28,
    **{index: 84 for index in range(1, 14)},
    14: 21,
    15: 105,
    16: 118,
    17: 105,
    18: 70,
    19: 70,
    20: 70,
    21: 70,
    22: 85,
    23: 85,
    24: 85,
}
DEFAULT_CONFIG = {
    "sectionNumber": "1",
    "sectionTitle": "Attributed Growth by Segment Template",
    "sheetName": "Attributed Growth",
    "attributedLabel": "Attributed Growth/Contribution",
    "bankOrder": DEFAULT_BANK_ORDER,
    "requireCompleteBankOrder": True,
    "useLogos": True,
    "logoDir": str(DEFAULT_LOGO_DIR),
    "logoFiles": DEFAULT_LOGO_FILES,
    "segments": [
        {"key": "CB", "label": "CB", "color": "#40484A", "textColor": "#FFFFFF"},
        {"key": "IB", "label": "IB", "color": "#001A52", "textColor": "#FFFFFF"},
        {"key": "CM", "label": "CM", "color": "#2D5967", "textColor": "#FFFFFF"},
        {"key": "WM", "label": "WM", "color": "#F28C00", "textColor": "#000000"},
        {"key": "CS_INS", "label": "CS + Ins.", "color": "#000000", "textColor": "#FFFFFF"},
    ],
    "banks": [
        {
            "bank": "RBC",
            "absoluteGrowthBn": 7.4,
            "totalGrowthPct": 12,
            "brandColor": "#0B3A78",
            "segments": {"CB": 5, "IB": 1, "CM": 3, "WM": 3, "CS_INS": 0},
        },
        {
            "bank": "BMO",
            "absoluteGrowthBn": 3.0,
            "totalGrowthPct": 9,
            "brandColor": "#0072A8",
            "segments": {"CB": 2, "IB": 1, "CM": 2, "WM": 2, "CS_INS": 2},
        },
        {
            "bank": "Scotia",
            "absoluteGrowthBn": 3.7,
            "totalGrowthPct": 11,
            "brandColor": "#D71920",
            "segments": {"CB": 1, "IB": 0, "CM": 3, "WM": 2, "CS_INS": 4},
        },
        {
            "bank": "CIBC",
            "absoluteGrowthBn": 3.6,
            "totalGrowthPct": 13,
            "brandColor": "#8B1020",
            "segments": {"CB": 5, "IB": 1, "CM": 6, "WM": 2, "CS_INS": -1},
        },
        {
            "bank": "National Bank",
            "shortName": "National",
            "absoluteGrowthBn": 1.7,
            "totalGrowthPct": 15,
            "attributed": True,
            "brandColor": "#C6002B",
            "segments": {"CB": 2, "IB": 1, "CM": 7, "WM": 2, "CS_INS": 3},
        },
        {
            "bank": "TD",
            "absoluteGrowthBn": 5.9,
            "totalGrowthPct": 12,
            "attributed": True,
            "brandColor": "#2E7D32",
            "segments": {"CB": 2, "IB": 2, "CM": 3, "WM": 3, "CS_INS": 2},
        },
    ],
}

BANK_ALIASES = {
    "rbc": "RBC",
    "ry": "RBC",
    "royal bank of canada": "RBC",
    "bmo": "BMO",
    "bank of montreal": "BMO",
    "scotia": "Scotia",
    "scotiabank": "Scotia",
    "bns": "Scotia",
    "bank of nova scotia": "Scotia",
    "cibc": "CIBC",
    "cm": "CIBC",
    "canadian imperial bank of commerce": "CIBC",
    "national": "National Bank",
    "national bank": "National Bank",
    "national bank of canada": "National Bank",
    "na": "National Bank",
    "nbc": "National Bank",
    "td": "TD",
    "td bank": "TD",
    "toronto dominion": "TD",
    "toronto dominion bank": "TD",
}


def normalize_text(value: Any) -> str:
    text = str(value or "").strip().lower().replace("&", "and")
    return re.sub(r"[^a-z0-9]+", " ", text).strip()


def canonical_bank_name(name: Any) -> str:
    raw = str(name or "").strip()
    return BANK_ALIASES.get(normalize_text(raw), raw)


def to_number(value: Any, fallback: float = 0) -> float:
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str) and value.strip():
        cleaned = re.sub(r"[%$,]", "", value, flags=re.I)
        cleaned = re.sub(r"\bBN\b", "", cleaned, flags=re.I).strip()
        try:
            return float(cleaned)
        except ValueError:
            return fallback
    return fallback


def to_bool(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    if isinstance(value, str):
        return value.strip().lower() in {"true", "yes", "y", "1"}
    return bool(value)


def row_value(row: dict[str, Any], keys: list[str]) -> Any:
    for key in keys:
        if key in row:
            return row[key]
    return None


def first_present(*values: Any) -> Any:
    for value in values:
        if value is not None:
            return value
    return None


def normalize_logo_files(logo_files: dict[str, str]) -> dict[str, str]:
    return {canonical_bank_name(bank): file_name for bank, file_name in logo_files.items()}


def normalize_bank_record(bank: dict[str, Any]) -> dict[str, Any]:
    canonical = canonical_bank_name(bank.get("bank"))
    defaults = next(
        (candidate for candidate in DEFAULT_CONFIG["banks"] if candidate["bank"] == canonical),
        {},
    )
    normalized = deepcopy(defaults)
    normalized.update(bank)
    normalized["bank"] = canonical
    normalized["absoluteGrowthBn"] = to_number(normalized.get("absoluteGrowthBn"))
    normalized["totalGrowthPct"] = to_number(normalized.get("totalGrowthPct"))
    normalized["shortName"] = bank.get("shortName", defaults.get("shortName"))
    normalized["brandColor"] = bank.get("brandColor", defaults.get("brandColor", "#111827"))
    normalized["segments"] = {
        **deepcopy(defaults.get("segments", {})),
        **{
            key: to_number(value)
            for key, value in deepcopy(bank.get("segments") or {}).items()
        },
    }
    return normalized


def normalize_data_rows(rows: list[dict[str, Any]], segments: list[dict[str, Any]]) -> list[dict[str, Any]]:
    normalized_rows = []
    for row in rows:
        bank_name = row_value(row, ["bank", "Bank"])
        if not bank_name:
            continue
        segment_values = {}
        for segment in segments:
            segment_values[segment["key"]] = to_number(
                first_present(
                    (row.get("segments") or {}).get(segment["key"]),
                    (row.get("segments") or {}).get(segment["label"]),
                    row_value(row, [segment["key"], segment["label"]]),
                )
            )
        normalized = {
            "bank": bank_name,
            "absoluteGrowthBn": to_number(
                row_value(row, ["absoluteGrowthBn", "Absolute Growth ($BN)", "Absolute Growth"])
            ),
            "totalGrowthPct": to_number(
                row_value(row, ["totalGrowthPct", "Reported YoY Growth", "YoY Growth"])
            ),
            "segments": segment_values,
        }
        for target, keys in {
            "shortName": ["shortName", "Short Name", "ShortName"],
            "brandColor": ["brandColor", "Brand Color"],
            "logoPath": ["logoPath", "Logo Path"],
            "logoFile": ["logoFile", "Logo File"],
        }.items():
            value = row_value(row, keys)
            if value not in (None, ""):
                normalized[target] = value
        attributed = row_value(row, ["attributed", "Attributed"])
        if attributed not in (None, ""):
            normalized["attributed"] = to_bool(attributed)
        normalized_rows.append(normalized)
    return normalized_rows


def order_banks(
    banks: list[dict[str, Any]], bank_order: list[str], require_complete: bool = True
) -> list[dict[str, Any]]:
    normalized_banks = [normalize_bank_record(bank) for bank in banks]
    order = [canonical_bank_name(bank) for bank in bank_order]
    order_set = set(order)
    banks_by_name = {bank["bank"]: bank for bank in normalized_banks}
    missing = [bank for bank in order if bank not in banks_by_name]
    unknown = [bank["bank"] for bank in normalized_banks if bank["bank"] not in order_set]
    if missing and require_complete:
        raise ValueError(f"Missing required bank rows: {', '.join(missing)}")
    if unknown:
        raise ValueError(
            f"Unknown bank rows: {', '.join(unknown)}. Add them to bankOrder before rendering."
        )
    if len(order) > len(CENTER_COLS):
        raise ValueError(f"The fixed chart layout supports up to {len(CENTER_COLS)} banks.")
    return [banks_by_name[name] for name in order if name in banks_by_name]


def prepare_config(parsed: dict[str, Any]) -> dict[str, Any]:
    parsed = deepcopy(parsed)
    default_segments = DEFAULT_CONFIG["segments"]
    segments = []
    for segment in parsed.get("segments", default_segments):
        defaults = next(
            (
                default_segment
                for default_segment in default_segments
                if default_segment["key"] == segment.get("key")
                or default_segment["label"] == segment.get("label")
            ),
            {},
        )
        merged = deepcopy(defaults)
        merged.update(segment)
        segments.append(merged)

    raw_banks = (
        normalize_data_rows(parsed["data"], segments)
        if "data" in parsed
        else deepcopy(parsed.get("banks", DEFAULT_CONFIG["banks"]))
    )
    bank_order = parsed.get("bankOrder", DEFAULT_CONFIG["bankOrder"])
    config = deepcopy(DEFAULT_CONFIG)
    config.update(parsed)
    config["segments"] = segments
    config["bankOrder"] = bank_order
    config["banks"] = order_banks(
        raw_banks,
        bank_order,
        parsed.get("requireCompleteBankOrder", DEFAULT_CONFIG["requireCompleteBankOrder"]),
    )
    config["logoFiles"] = {
        **DEFAULT_LOGO_FILES,
        **normalize_logo_files(parsed.get("logoFiles", {})),
    }
    return config


def load_config_from_input_workbook(input_path: Path) -> dict[str, Any]:
    try:
        from openpyxl import load_workbook
    except ModuleNotFoundError as exc:
        raise SystemExit(
            "Missing dependency: openpyxl. Run: .venv/bin/pip install -r requirements.txt"
        ) from exc

    workbook = load_workbook(input_path, data_only=True)
    if INPUT_LAYOUT["sheet_name"] not in workbook.sheetnames:
        raise ValueError(f'Could not find sheet "{INPUT_LAYOUT["sheet_name"]}" in {input_path}')
    sheet = workbook[INPUT_LAYOUT["sheet_name"]]

    def cell(address: str) -> Any:
        value = sheet[address].value
        return value.strip() if isinstance(value, str) else value

    config: dict[str, Any] = {
        "sectionNumber": cell("B1") or DEFAULT_CONFIG["sectionNumber"],
        "sectionTitle": cell("B2") or DEFAULT_CONFIG["sectionTitle"],
    }
    segments = []
    for offset in range(INPUT_LAYOUT["segment_max_rows"]):
        row = INPUT_LAYOUT["segment_start_row"] + offset
        key = cell(f"A{row}")
        if key:
            segments.append({"key": str(key).strip(), "label": cell(f"B{row}") or str(key).strip()})
    if segments:
        config["segments"] = segments
    active_segments = config.get("segments", DEFAULT_CONFIG["segments"])

    rows = []
    for offset in range(INPUT_LAYOUT["data_max_rows"]):
        row_number = INPUT_LAYOUT["data_start_row"] + offset
        bank = cell(f"A{row_number}")
        if not bank:
            continue
        row = {
            "bank": bank,
            "absoluteGrowthBn": cell(f"B{row_number}"),
            "totalGrowthPct": cell(f"C{row_number}"),
        }
        for index, segment in enumerate(active_segments):
            col = xlsxwriter.utility.xl_col_to_name(INPUT_LAYOUT["data_segment_start_col"] + index)
            row[segment["key"]] = cell(f"{col}{row_number}")
        rows.append(row)

    return {**config, "bankOrder": [row["bank"] for row in rows], "data": rows}


def load_config(input_path: Path | None) -> dict[str, Any]:
    if not input_path:
        return prepare_config({})
    if input_path.suffix.lower() == ".xlsx":
        return prepare_config(load_config_from_input_workbook(input_path))
    return prepare_config(json.loads(input_path.read_text()))


def safe_sheet_name(config: dict[str, Any]) -> str:
    raw = str(config.get("sheetName") or config.get("sectionTitle") or "Segment Growth")
    cleaned = re.sub(r"[\[\]:*?/\\]", " ", raw)[:31].strip()
    return cleaned or "Segment Growth"


def section_header(config: dict[str, Any]) -> str:
    return (
        f'{config["sectionNumber"]}. {config["sectionTitle"]}'
        if config.get("sectionNumber")
        else config["sectionTitle"]
    )


def segment_value(bank: dict[str, Any], segment_key: str) -> float:
    value = bank.get("segments", {}).get(segment_key, 0)
    return float(value) if isinstance(value, (int, float)) else 0


def abs_growth_label(value: float) -> str:
    sign = "+" if value >= 0 else "-"
    return f"{sign}{abs(value):.1f}BN"


def pct_label(value: float) -> str:
    return f"{round(value)}%"


def row_height_px(row_index: int) -> int:
    if row_index == 0:
        return 36
    if row_index == 1:
        return 24
    if row_index == 2:
        return 10
    return 20


def column_width_px(column_index: int) -> int:
    return COLUMN_WIDTHS_PX.get(column_index, 64)


def row_top_px(row_index: int) -> int:
    return sum(row_height_px(row) for row in range(row_index))


def column_left_px(column_index: int) -> int:
    return sum(column_width_px(col) for col in range(column_index))


def category_center_px(bank_index: int) -> float:
    column = CENTER_COLS[bank_index]
    return column_left_px(column) + column_width_px(column) / 2


def pixel_to_row_anchor(y_px: float) -> tuple[int, int]:
    row = 0
    remaining = max(0, float(y_px))
    while remaining >= row_height_px(row):
        remaining -= row_height_px(row)
        row += 1
    return row, round(remaining)


def pixel_to_col_anchor(x_px: float) -> tuple[int, int]:
    col = 0
    remaining = max(0, float(x_px))
    while remaining >= column_width_px(col):
        remaining -= column_width_px(col)
        col += 1
    return col, round(remaining)


def textbox_at_px(
    worksheet: xlsxwriter.worksheet.Worksheet,
    x_px: float,
    y_px: float,
    width_px: float,
    height_px: float,
    text: str = "",
    *,
    fill_color: str | None = "#FFFFFF",
    line_color: str | None = None,
    font_color: str = "#000000",
    font_size: int = 11,
    bold: bool = True,
    object_position: int = 2,
) -> None:
    row, y_offset = pixel_to_row_anchor(y_px)
    col, x_offset = pixel_to_col_anchor(x_px)
    options: dict[str, Any] = {
        "x_offset": x_offset,
        "y_offset": y_offset,
        "width": max(1, round(width_px)),
        "height": max(1, round(height_px)),
        "object_position": object_position,
        "align": {"horizontal": "center", "vertical": "middle"},
        "font": {"color": font_color, "size": font_size, "bold": bold},
        "margin": 0,
    }
    if fill_color is None:
        options["fill"] = {"none": True}
    else:
        options["fill"] = {"color": fill_color}
    if line_color is None:
        options["line"] = {"none": True}
    else:
        options["line"] = {"color": line_color}
    worksheet.insert_textbox(row, col, text, options)


def resolve_logo_path(bank: dict[str, Any], config: dict[str, Any]) -> Path | None:
    if config.get("useLogos") is False:
        return None
    if bank.get("logoPath"):
        explicit = Path(bank["logoPath"]).expanduser().resolve()
        return explicit if explicit.is_file() else None
    logo_file = bank.get("logoFile") or config.get("logoFiles", {}).get(bank["bank"])
    logo_dir = config.get("logoDir")
    if not logo_file or not logo_dir:
        return None
    candidate = Path(logo_dir).expanduser().resolve() / logo_file
    return candidate if candidate.is_file() else None


def png_dimensions(file_path: Path) -> tuple[int, int] | None:
    data = file_path.read_bytes()
    if len(data) >= 24 and data[:8] == b"\x89PNG\r\n\x1a\n":
        return int.from_bytes(data[16:20], "big"), int.from_bytes(data[20:24], "big")
    return None


def normalized_logo_path(file_path: Path, temp_dir: Path) -> Path:
    if Image is None:
        return file_path

    image = Image.open(file_path).convert("RGBA")
    alpha_bbox = image.getchannel("A").getbbox()
    bbox = alpha_bbox or image.getbbox()
    if bbox:
        image = image.crop(bbox)

    output_path = temp_dir / file_path.name
    image.save(output_path)
    return output_path


def logo_layout(file_path: Path) -> tuple[int, int]:
    max_width, max_height = 126, 40
    dimensions = png_dimensions(file_path)
    if not dimensions:
        return 96, 34
    width, height = dimensions
    scale = min(max_width / width, max_height / height)
    return round(width * scale), round(height * scale)


def logo_position(logo_path: Path, bank_index: int) -> tuple[int, int, int, int, int, int]:
    center_x = category_center_px(bank_index)
    width_px, height_px = logo_layout(logo_path)
    total_label_height_px = 44
    x_px = center_x - width_px / 2
    col, x_offset = pixel_to_col_anchor(x_px)
    y_offset = max(0, round((total_label_height_px - height_px) / 2))
    return LAYOUT["logo_top_row"], col, x_offset, y_offset, width_px, height_px


def make_formats(workbook: xlsxwriter.Workbook) -> dict[str, Any]:
    return {
        "title": workbook.add_format(
            {
                "bold": True,
                "font_size": 20,
                "font_color": "#FFFFFF",
                "bg_color": "#071734",
                "valign": "vcenter",
            }
        ),
        "top_abs": workbook.add_format({"bold": True, "font_size": 17, "align": "center"}),
        "top_pct": workbook.add_format({"bold": True, "font_size": 15, "align": "center"}),
        "attributed": workbook.add_format(
            {
                "bold": True,
                "font_size": 15,
                "font_color": "#111827",
                "bg_color": "#BDBDBD",
                "align": "center",
                "valign": "vcenter",
            }
        ),
        "source_header": workbook.add_format(
            {"bold": True, "font_size": 9, "font_color": "#FFFFFF", "bg_color": "#0D6786"}
        ),
        "source": workbook.add_format({"font_size": 9}),
        "source_bn": workbook.add_format({"font_size": 9, "num_format": '+0.0"BN";-0.0"BN";0.0"BN"'}),
        "source_pct": workbook.add_format({"font_size": 9, "num_format": '0"%"'}),
        "input_header": workbook.add_format(
            {"bold": True, "font_color": "#FFFFFF", "bg_color": "#0D6786"}
        ),
        "input_label": workbook.add_format({"bold": True, "font_color": "#111827", "bg_color": "#D9EAF7"}),
        "input_bn": workbook.add_format({"num_format": '+0.0"BN";-0.0"BN";0.0"BN"'}),
        "input_pct": workbook.add_format({"num_format": '0"%"'}),
    }


def setup_common_sheet(worksheet: xlsxwriter.worksheet.Worksheet) -> None:
    worksheet.hide_gridlines(2)
    worksheet.set_landscape()
    worksheet.fit_to_pages(1, 1)
    worksheet.set_margins(0.3, 0.3, 0.3, 0.3)
    worksheet.print_area(0, 0, 31, 13)
    worksheet.center_horizontally()
    for col in range(25):
        worksheet.set_column_pixels(col, col, COLUMN_WIDTHS_PX.get(col, 64))
    for row in range(32):
        worksheet.set_row_pixels(row, row_height_px(row))


def write_source_table(
    workbook: xlsxwriter.Workbook,
    worksheet: xlsxwriter.worksheet.Worksheet,
    config: dict[str, Any],
    formats: dict[str, Any],
) -> tuple[int, int, int, int]:
    header = [
        "Bank",
        "Absolute Growth ($BN)",
        "Reported YoY Growth",
        *[segment["label"] for segment in config["segments"]],
        "Attributed",
    ]
    rows = [
        [
            bank["bank"],
            "" if config.get("templateMode") else bank["absoluteGrowthBn"],
            "" if config.get("templateMode") else bank["totalGrowthPct"],
            *[
                "" if config.get("templateMode") else segment_value(bank, segment["key"])
                for segment in config["segments"]
            ],
            "Yes" if bank.get("attributed") else "",
        ]
        for bank in config["banks"]
    ]
    start_row = LAYOUT["source_table_start_row"]
    start_col = 15
    end_row = start_row + len(rows)
    end_col = start_col + len(header) - 1

    worksheet.add_table(
        start_row,
        start_col,
        end_row,
        end_col,
        {
            "name": "AttributedGrowthBySegmentSource",
            "style": "Table Style Medium 2",
            "data": rows,
            "columns": [{"header": title} for title in header],
        },
    )
    for col_offset, title in enumerate(header):
        worksheet.write(start_row, start_col + col_offset, title, formats["source_header"])
    for row_offset, row in enumerate(rows, start=1):
        for col_offset, value in enumerate(row):
            fmt = formats["source"]
            if col_offset == 1:
                fmt = formats["source_bn"]
            elif 2 <= col_offset <= 2 + len(config["segments"]):
                fmt = formats["source_pct"]
            worksheet.write(start_row + row_offset, start_col + col_offset, value, fmt)
    return start_row, start_col, end_row, end_col


def add_native_chart(
    workbook: xlsxwriter.Workbook,
    worksheet: xlsxwriter.worksheet.Worksheet,
    sheet_name: str,
    config: dict[str, Any],
    source_bounds: tuple[int, int, int, int],
) -> None:
    start_row, start_col, end_row, _end_col = source_bounds
    chart = workbook.add_chart({"type": "column", "subtype": "stacked"})
    chart.show_hidden_data()
    category_range = [sheet_name, start_row + 1, start_col, end_row, start_col]
    for index, segment in enumerate(config["segments"]):
        col = start_col + 3 + index
        chart.add_series(
            {
                "name": [sheet_name, start_row, col],
                "categories": category_range,
                "values": [sheet_name, start_row + 1, col, end_row, col],
                "fill": {"color": segment["color"]},
                "border": {"color": "#FFFFFF"},
                "data_labels": {
                    "value": True,
                    "position": "center",
                    "num_format": '0"%"',
                    "font": {"color": segment.get("textColor", "#FFFFFF"), "size": 12},
                },
            }
        )
    extents = segment_stack_extents(config)
    axis_max = max(1, extents["positive"] * CHART_GEOMETRY["axis_max_padding"])
    axis_min = -max(1.5, extents["negative"] * 1.5) if extents["negative"] > 0 else 0
    chart.set_title({"none": True})
    chart.set_legend({"none": True})
    chart.set_chartarea({"fill": {"color": "#FFFFFF"}, "border": {"none": True}})
    chart.set_plotarea({"fill": {"color": "#FFFFFF"}, "border": {"none": True}})
    chart.set_x_axis(
        {
            "visible": False,
            "major_gridlines": {"visible": False},
            "line": {"none": True},
            "num_font": {"color": "#FFFFFF"},
        }
    )
    chart.set_y_axis(
        {
            "visible": False,
            "min": axis_min,
            "max": axis_max,
            "major_unit": 5,
            "major_gridlines": {"visible": False},
            "line": {"none": True},
        }
    )
    chart.set_size({"width": 980, "height": 470})
    chart.series_gap_1 = 70
    chart.series_gap_2 = 70
    chart.series_overlap_1 = 100
    chart.series_overlap_2 = 100
    worksheet.insert_chart(LAYOUT["chart_top_row"], LAYOUT["chart_top_col"], chart)


def chart_axis_geometry(config: dict[str, Any]) -> dict[str, float]:
    extents = segment_stack_extents(config)
    axis_max = max(1, extents["positive"] * CHART_GEOMETRY["axis_max_padding"])
    axis_min = -max(1.5, extents["negative"] * 1.5) if extents["negative"] > 0 else 0
    plot_top = row_top_px(CHART_GEOMETRY["top_row"])
    plot_bottom = row_top_px(CHART_GEOMETRY["bottom_row"])
    unit_px = (plot_bottom - plot_top) / (axis_max - axis_min)
    baseline = plot_top + axis_max * unit_px
    first_center = category_center_px(0)
    last_center = category_center_px(len(config["banks"]) - 1)
    bar_width = CHART_GEOMETRY["bar_width_px"]
    return {
        "plot_top": plot_top,
        "plot_bottom": plot_bottom,
        "unit_px": unit_px,
        "baseline": baseline,
        "line_left": first_center - bar_width / 2 - 42,
        "line_right": last_center + bar_width / 2 + 42,
    }


def stagger_label_positions(labels: list[dict[str, Any]]) -> None:
    labels.sort(key=lambda label: label["center_y"])
    cluster: list[dict[str, Any]] = []
    min_gap = CHART_GEOMETRY["label_height_px"] + 2

    def apply_cluster_offsets(cluster_labels: list[dict[str, Any]]) -> None:
        if len(cluster_labels) <= 1:
            return
        offsets = [-22, 22, 0, -34, 34]
        for index, label in enumerate(cluster_labels):
            label["x_shift"] = offsets[index % len(offsets)]

    for label in labels:
        if not cluster:
            cluster = [label]
            continue
        if label["center_y"] - cluster[-1]["center_y"] < min_gap:
            cluster.append(label)
        else:
            apply_cluster_offsets(cluster)
            cluster = [label]
    apply_cluster_offsets(cluster)


def draw_manual_chart(
    worksheet: xlsxwriter.worksheet.Worksheet, config: dict[str, Any]
) -> None:
    geometry = chart_axis_geometry(config)
    bar_width = CHART_GEOMETRY["bar_width_px"]
    label_width = CHART_GEOMETRY["label_width_px"]
    label_height = CHART_GEOMETRY["label_height_px"]
    bar_overlap = CHART_GEOMETRY["bar_overlap_px"]
    zero_marker_height = CHART_GEOMETRY["zero_marker_height_px"]
    all_labels: list[dict[str, Any]] = []

    if config.get("templateMode"):
        textbox_at_px(
            worksheet,
            geometry["line_left"],
            geometry["baseline"] - 1,
            geometry["line_right"] - geometry["line_left"],
            2,
            fill_color=CHART_GEOMETRY["zero_line_color"],
        )
        return

    for bank_index, bank in enumerate(config["banks"]):
        if bank_index >= len(CENTER_COLS):
            continue
        center_x = category_center_px(bank_index)
        left_x = center_x - bar_width / 2
        positive_stack = 0.0
        negative_stack = 0.0
        bank_labels: list[dict[str, Any]] = []
        zero_marker_counts: dict[int, int] = {}

        for segment in config["segments"]:
            value = segment_value(bank, segment["key"])
            fill = segment["color"]
            text_color = segment.get("textColor", "#FFFFFF")

            if value == 0:
                marker_base = (
                    geometry["baseline"] - positive_stack * geometry["unit_px"]
                    if positive_stack > 0
                    else geometry["baseline"] + negative_stack * geometry["unit_px"]
                )
                marker_key = round(marker_base)
                marker_count = zero_marker_counts.get(marker_key, 0)
                zero_marker_counts[marker_key] = marker_count + 1
                marker_y = marker_base + marker_count * (zero_marker_height + 1)
                textbox_at_px(
                    worksheet,
                    left_x,
                    marker_y - zero_marker_height / 2,
                    bar_width,
                    zero_marker_height,
                    fill_color=fill,
                    line_color=fill,
                )
                bank_labels.append(
                    {
                        "text": "0%",
                        "fill": fill,
                        "text_color": text_color,
                        "center_x": center_x,
                        "center_y": marker_y,
                        "x_shift": 0,
                    }
                )
                continue

            height = max(1, abs(value) * geometry["unit_px"])
            top_y = (
                geometry["baseline"] - (positive_stack + value) * geometry["unit_px"]
                if value > 0
                else geometry["baseline"] + negative_stack * geometry["unit_px"]
            )
            draw_height = height + bar_overlap
            textbox_at_px(
                worksheet,
                left_x,
                top_y,
                bar_width,
                draw_height,
                fill_color=fill,
                line_color=fill,
            )
            rounded = round(value)
            if rounded != 0:
                bank_labels.append(
                    {
                        "text": pct_label(value),
                        "fill": fill,
                        "text_color": text_color,
                        "center_x": center_x,
                        "center_y": top_y + height / 2,
                        "x_shift": 0,
                    }
                )
            if value > 0:
                positive_stack += value
            else:
                negative_stack += abs(value)

        stagger_label_positions(bank_labels)
        all_labels.extend(bank_labels)

    textbox_at_px(
        worksheet,
        geometry["line_left"],
        geometry["baseline"] - 1,
        geometry["line_right"] - geometry["line_left"],
        2,
        fill_color=CHART_GEOMETRY["zero_line_color"],
    )

    for label in all_labels:
        textbox_at_px(
            worksheet,
            label["center_x"] - label_width / 2 + label["x_shift"],
            label["center_y"] - label_height / 2,
            label_width,
            label_height,
            label["text"],
            fill_color=label["fill"],
            line_color=label["fill"],
            font_color=label["text_color"],
            font_size=12,
            bold=True,
        )


def segment_stack_extents(config: dict[str, Any]) -> dict[str, float]:
    positive_max = 0.0
    negative_max = 0.0
    for bank in config["banks"]:
        positive = 0.0
        negative = 0.0
        for segment in config["segments"]:
            value = segment_value(bank, segment["key"])
            if value > 0:
                positive += value
            elif value < 0:
                negative += abs(value)
        positive_max = max(positive_max, positive)
        negative_max = max(negative_max, negative)
    return {"positive": positive_max, "negative": negative_max}


def write_attributed_growth_workbook(config: dict[str, Any], output_path: Path) -> None:
    sheet_name = safe_sheet_name(config)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = xlsxwriter.Workbook(output_path)
    worksheet = workbook.add_worksheet(sheet_name)
    formats = make_formats(workbook)
    setup_common_sheet(worksheet)

    worksheet.merge_range(0, 0, 0, 24, section_header(config), formats["title"])

    legend_cells = [1, 2, 3, 4, 5]
    for col, segment in zip(legend_cells, config["segments"]):
        legend_format = workbook.add_format({"font_size": 14, "font_color": segment["color"]})
        worksheet.write(1, col, f"{chr(9632)} {segment['label']}", legend_format)

    attributed = [(bank, index) for index, bank in enumerate(config["banks"]) if bank.get("attributed")]
    if attributed:
        first = CENTER_COLS[attributed[0][1]]
        last = CENTER_COLS[attributed[-1][1]] + 1
        worksheet.merge_range(1, first, 1, last, config["attributedLabel"], formats["attributed"])

    for index, bank in enumerate(config["banks"]):
        if index >= len(CENTER_COLS):
            continue
        col = CENTER_COLS[index]
        worksheet.write(
            LAYOUT["absolute_label_row"],
            col,
            "" if config.get("templateMode") else abs_growth_label(bank["absoluteGrowthBn"]),
            formats["top_abs"],
        )
        worksheet.write(
            LAYOUT["total_label_row"],
            col,
            "" if config.get("templateMode") else pct_label(bank["totalGrowthPct"]),
            formats["top_pct"],
        )

    source_bounds = write_source_table(workbook, worksheet, config, formats)

    native_chart_sheet = workbook.add_worksheet("_Native Chart")
    native_chart_sheet.hide()
    add_native_chart(workbook, native_chart_sheet, sheet_name, config, source_bounds)
    draw_manual_chart(worksheet, config)

    with tempfile.TemporaryDirectory() as temp_dir_name:
        temp_dir = Path(temp_dir_name)
        for index, bank in enumerate(config["banks"]):
            if index >= len(CENTER_COLS):
                continue
            logo_path = resolve_logo_path(bank, config)
            col = CENTER_COLS[index]
            if logo_path:
                display_logo = normalized_logo_path(logo_path, temp_dir)
                row, image_col, x_offset, y_offset, width_px, height_px = logo_position(
                    display_logo, index
                )
                dimensions = png_dimensions(display_logo) or (width_px, height_px)
                worksheet.insert_image(
                    row,
                    image_col,
                    str(display_logo),
                    {
                        "x_offset": x_offset,
                        "y_offset": y_offset,
                        "x_scale": width_px / dimensions[0],
                        "y_scale": height_px / dimensions[1],
                        "object_position": 2,
                    },
                )
            else:
                brand_format = workbook.add_format(
                    {
                        "bold": True,
                        "font_size": 14 if bank["bank"] == "National Bank" else 18,
                        "font_color": bank.get("brandColor", "#111827"),
                        "align": "center",
                        "valign": "vcenter",
                    }
                )
                worksheet.merge_range(
                    LAYOUT["logo_top_row"],
                    col,
                    LAYOUT["logo_bottom_row"],
                    col,
                    bank.get("shortName") or bank["bank"],
                    brand_format,
                )

        workbook.close()


def write_input_template_workbook(config: dict[str, Any], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = xlsxwriter.Workbook(output_path)
    worksheet = workbook.add_worksheet(INPUT_LAYOUT["sheet_name"])
    formats = make_formats(workbook)
    for col, width in enumerate([140, 126, 126, 70, 70, 70, 70, 70]):
        worksheet.set_column_pixels(col, col, width)

    worksheet.write_row(0, 0, ["Section Number", config["sectionNumber"]])
    worksheet.write_row(1, 0, ["Section Title", config["sectionTitle"]])
    worksheet.write(0, 0, "Section Number", formats["input_label"])
    worksheet.write(1, 0, "Section Title", formats["input_label"])

    segment_header_row = INPUT_LAYOUT["segment_start_row"] - 2
    segment_rows = [[segment["key"], segment["label"]] for segment in config["segments"]]
    worksheet.add_table(
        segment_header_row,
        0,
        segment_header_row + len(segment_rows),
        1,
        {
            "name": "AttributedGrowthSegmentsInput",
            "style": "Table Style Medium 2",
            "data": segment_rows,
            "columns": [{"header": "Segment Key"}, {"header": "Segment Label"}],
        },
    )
    worksheet.write_row(segment_header_row, 0, ["Segment Key", "Segment Label"], formats["input_header"])

    data_headers = [
        "Bank",
        "Absolute Growth ($BN)",
        "Reported YoY Growth",
        *[segment["key"] for segment in config["segments"]],
    ]
    data_rows = [
        [
            bank["bank"],
            bank["absoluteGrowthBn"],
            bank["totalGrowthPct"],
            *[segment_value(bank, segment["key"]) for segment in config["segments"]],
        ]
        for bank in config["banks"]
    ]
    data_header_row = INPUT_LAYOUT["data_start_row"] - 2
    worksheet.add_table(
        data_header_row,
        0,
        data_header_row + len(data_rows),
        len(data_headers) - 1,
        {
            "name": "AttributedGrowthBankDataInput",
            "style": "Table Style Medium 2",
            "data": data_rows,
            "columns": [{"header": header} for header in data_headers],
        },
    )
    worksheet.write_row(data_header_row, 0, data_headers, formats["input_header"])
    for row_offset, row in enumerate(data_rows, start=1):
        for col_offset, value in enumerate(row):
            fmt = None
            if col_offset == 1:
                fmt = formats["input_bn"]
            elif col_offset >= 2:
                fmt = formats["input_pct"]
            worksheet.write(data_header_row + row_offset, col_offset, value, fmt)
    workbook.close()


def blank_template_config() -> dict[str, Any]:
    segments = deepcopy(DEFAULT_CONFIG["segments"])
    empty_segments = {segment["key"]: 0 for segment in segments}
    return prepare_config(
        {
            "banks": [
                {
                    **deepcopy(bank),
                    "absoluteGrowthBn": 0,
                    "totalGrowthPct": 0,
                    "segments": deepcopy(empty_segments),
                }
                for bank in DEFAULT_CONFIG["banks"]
            ],
            "segments": segments,
            "templateMode": True,
        }
    )


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Build attributed growth by segment template.")
    parser.add_argument("--input", type=Path)
    parser.add_argument("--output", type=Path)
    parser.add_argument("--create-input-template", nargs="?", const=True)
    parser.add_argument("--create-blank-template", nargs="?", const=True)
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    if args.create_input_template:
        output = (
            Path(args.create_input_template)
            if isinstance(args.create_input_template, str)
            else DEFAULT_INPUT_PATH
        )
        write_input_template_workbook(prepare_config({}), output.resolve())
        print(output.resolve())
        return

    output = (
        Path(args.create_blank_template)
        if isinstance(args.create_blank_template, str)
        else args.output
        or DEFAULT_OUTPUT_PATH
    )
    input_path = args.input or (None if args.create_blank_template else DEFAULT_INPUT_PATH)
    config = blank_template_config() if args.create_blank_template else load_config(input_path)
    write_attributed_growth_workbook(config, output.resolve())
    print(output.resolve())


if __name__ == "__main__":
    main()
